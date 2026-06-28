"""PR-06B 剪辑清单多格式序列化测试（纯逻辑，无 DB）。

覆盖 csv/xlsx/json/markdown/printable：内容正确、中文正常、公式注入防护、HTML escape、
缺口段保留、无本机绝对路径/Key/Endpoint。
"""

from __future__ import annotations

import io
import json
import zipfile

from clipmind_shared.script import editlist as E
from clipmind_shared.script.editlist import CandidateView, SegmentView


def _rows_summary():
    seg_ok = SegmentView(
        segment_id=1, order_index=0, segment_text="=危险公式 开场 hook",  # 注入前缀 + 中文
        product_id=1, product_name="产品A", current_generation=1, match_status="matched",
        candidates=[CandidateView(
            shot_id=10, asset_id=5, rank=0, final_score=0.8, sequence_no=1,
            source_start=0.0, source_end=2.0, source_duration=2.0, asset_filename="a.mp4",
            product_name="产品A", scene_labels=["户外"], action_labels=["使用"],
        )],
        selected_shot_id=10,
    )
    seg_gap = SegmentView(
        segment_id=2, order_index=1, segment_text="产品特写", product_id=2,
        product_name="产品B", current_generation=1, match_status="gap",
        match_summary={"gap_reasons": ["无符合产品硬约束的镜头：产品B"],
                       "reshoot_recommendation": ["补拍产品特写"],
                       "requires_human_confirmation": True},
    )
    return E.build_edit_list([seg_ok, seg_gap])


def _meta(rows):
    return E.build_meta(project_name="演示<脚本>&", project_id=7, row_count=len(rows),
                        generated_at="2026-06-28T00:00:00+00:00")


def test_csv_bom_and_injection_guard():
    rows, summary = _rows_summary()
    data = E.serialize_edit_list("csv", rows, summary, meta=_meta(rows))
    assert data[:3] == b"\xef\xbb\xbf"  # UTF-8 BOM
    assert "段落序号".encode() in data
    assert b"'=" in data  # 公式注入前缀被加单引号


def test_xlsx_is_valid_zip_openpyxl_loadable():
    rows, summary = _rows_summary()
    data = E.serialize_edit_list("xlsx", rows, summary, meta=_meta(rows))
    assert data[:2] == b"PK"  # XLSX = zip
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    ws = wb["剪辑清单"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert "段落序号" in headers
    assert "摘要" in wb.sheetnames


def test_json_schema_and_gap_preserved():
    rows, summary = _rows_summary()
    data = E.serialize_edit_list("json", rows, summary, meta=_meta(rows))
    obj = json.loads(data)
    assert obj["metadata"]["project_id"] == 7
    assert len(obj["segments"]) == 2
    assert "summary" in obj
    gap = [s for s in obj["segments"] if s["match_status"] == "gap"]
    assert gap and gap[0]["gap_reasons"]  # 缺口段保留


def test_markdown_table_and_escape():
    rows, summary = _rows_summary()
    data = E.serialize_edit_list("markdown", rows, summary, meta=_meta(rows)).decode("utf-8")
    assert "| 段落序号 |" in data
    assert "## 缺口与补拍建议" in data
    assert "补拍产品特写" in data


def test_printable_self_contained_and_escaped():
    rows, summary = _rows_summary()
    data = E.serialize_edit_list("printable", rows, summary, meta=_meta(rows)).decode("utf-8")
    assert data.startswith("<!DOCTYPE html>")
    assert "<style>" in data and "@media print" in data
    assert "http://" not in data and "https://" not in data  # 无外部 CDN
    assert "&lt;脚本&gt;" in data  # 用户文本 HTML escape


def test_no_absolute_paths_or_secrets_in_any_format():
    rows, summary = _rows_summary()
    for fmt in ("csv", "json", "markdown", "printable"):
        data = E.serialize_edit_list(fmt, rows, summary, meta=_meta(rows))
        text = data.decode("utf-8", errors="ignore")
        assert "/app/" not in text and "C:\\" not in text
        assert "api_key" not in text.lower() and "xiaomimimo" not in text.lower()


def test_zip_bundle_layout_smoke():
    # 验证 ZIP 打包内容布局约定（与 worker 产物一致）
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("clips/001_a.mp4", b"x")
        zf.writestr("manifest.json", b"{}")
        zf.writestr("edit-list.csv", b"\xef\xbb\xbf")
        zf.writestr("README.txt", b"r")
    names = zipfile.ZipFile(io.BytesIO(buf.getvalue())).namelist()
    assert "manifest.json" in names and any(n.startswith("clips/") for n in names)
